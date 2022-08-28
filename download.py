from selenium import webdriver
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import os
import time
from multiprocessing import Process


def download(driver, Scode):
    time.sleep(1)
    # 输入股票代码，分两次输入才能使其弹出搜索结果
    driver.find_element_by_xpath('//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[1]/input').clear()
    driver.find_element_by_xpath('//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[1]/input').send_keys(Scode[0:3])
    # time.sleep(1)
    driver.find_element_by_xpath('//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[1]/input').send_keys(Scode[3:6])
    time.sleep(1)
    # 如果弹出了搜索结果，点击弹出的结果，下载数据
    try:
        driver.find_element_by_xpath('//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[1]/div').click()
        time.sleep(1.5)
        driver.find_element_by_xpath('//*[@id="marketXiazai"]').click()
        time.sleep(1)
    except:
        print(Scode + '不存在')
        return 0


def load_Scode(file_path, num):
    Scode = []
    Coname = []
    major = load_workbook(file_path)  # 主表格路径
    major_sheet = major[major.sheetnames[num - 1]]  # 通过num控制读取的表单
    # major_sheet = major['Sheet1']
    # 载入需要搜索的股票代码
    for i in range(2, 1600):
        if major_sheet['A' + str(i)].value is not None:
            Scode.append(str(major_sheet['A' + str(i)].value))
            Coname.append(str(major_sheet['B' + str(i)].value))
        else:
            return Scode, Coname


def split(list, sp_num):
    sp_dict = {}
    clip_num = len(list) // sp_num + 1
    for i in range(0, sp_num):
        if i == sp_num - 1:
            sp_dict[i] = list[i * clip_num:]
        else:
            sp_dict[i] = list[i * clip_num: (i + 1) * clip_num]
    return sp_dict


def bug(Scode):
    # start_time = '2020-01-03'
    # end_time = '2020-02-11'

    start_time = '2019-01-01'
    end_time = '2019-12-31'

    pinlv = '每日'

    # 设置文件的下载路径
    options = webdriver.ChromeOptions()
    # 修改主机的IP防止下载次数过多导致被拒绝访问
    options.add_argument("--proxy-server=http://112.6.117.178:8085")
    prefs = {'profile.default_content_settings.popups': 0,
             'download.default_directory': r'C:\Users\65487\Desktop\股票代码\股票代码\19\中小板\\'}
    options.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome(r'D:\code\bug\chromedriver.exe', chrome_options=options)
    driver.get('http://webapi.cninfo.com.cn/#/marketData')

    # 选择查询的开始时间与结束时间
    driver.find_element_by_xpath('//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[2]/div/input').click()
    time.sleep(1)
    driver.find_element_by_xpath(
        '//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[2]/div/div/div[2]/input[1]').clear()
    driver.find_element_by_xpath(
        '//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[2]/div/div/div[2]/input[2]').clear()
    driver.find_element_by_xpath(
        '//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[2]/div/div/div[2]/input[1]').send_keys(start_time)
    driver.find_element_by_xpath(
        '//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[2]/div/div/div[2]/input[2]').send_keys(end_time)
    driver.find_element_by_xpath(
        '//*[@id="root"]/div/div[3]/div/div/div[1]/div/div/div[2]/div/div/div[4]/button[1]').click()

    select = Select(driver.find_element_by_id('seee1'))
    if pinlv == '每日':
        index = 0
    elif pinlv == '每周':
        index = 1
    elif pinlv == '每年':
        index = 2
    select.select_by_index(index)

    time.sleep(30)

    for code in Scode:
        download(driver, code)


if __name__ == '__main__':
    Scode, Coname = load_Scode(r'C:\Users\65487\Desktop\股票代码\股票代码\中小板.xlsx', 1)
    # bug(Scode)
    sp_dict = split(Scode, 10)
    print(sp_dict)
    p0 = Process(target=bug, args=(sp_dict[0],))
    p1 = Process(target=bug, args=(sp_dict[1],))
    p2 = Process(target=bug, args=(sp_dict[2],))
    p3 = Process(target=bug, args=(sp_dict[3],))
    p4 = Process(target=bug, args=(sp_dict[4],))
    p5 = Process(target=bug, args=(sp_dict[5],))
    p6 = Process(target=bug, args=(sp_dict[6],))
    p7 = Process(target=bug, args=(sp_dict[7],))
    p8 = Process(target=bug, args=(sp_dict[8],))
    p9 = Process(target=bug, args=(sp_dict[9],))

    p0.start()
    p1.start()
    p2.start()
    p3.start()
    p4.start()
    p5.start()
    p6.start()
    p7.start()
    p8.start()
    p9.start()

    p0.join()
    p1.join()
    p2.join()
    p3.join()
    p4.join()
    p5.join()
    p6.join()
    p7.join()
    p8.join()
    p9.join()
