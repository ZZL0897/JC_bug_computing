import numpy as np
from openpyxl import load_workbook, Workbook
import os
from pathlib import Path
import csv


def Regression(x, y):
    # 计算x, y的回归函数a, b的值
    res = np.polyfit(x, y, deg=1)
    a = res[0]
    b = res[1]
    return a, b


def load_price(file_path, start, end, line='F'):
    price = []
    file_path = Path(file_path)
    major = load_workbook(file_path)
    # file_name = file_path.stem
    major_sheet = major[major.sheetnames[0]]
    # 字母根据数据所在列进行修改（两个都要改），range的范围根据数据数量进行修改
    for i in range(start, end):
        if major_sheet[line + str(i)].value is not None:
            price.append(major_sheet[line + str(i)].value)
        else:
            break
    price.reverse()
    return price


def zhangdie(price):
    zhangdie = []
    # 把空缺日期的数据补全为上一天的，防止除法除以0出错
    for i in range(0, len(price)):
        if price[i] == 0:
            price[i] = price[i - 1]
    # 根据公式计算涨跌幅
    for i in range(0, len(price) - 1):
        y = (price[i + 1] - price[i]) / price[i]
        zhangdie.append(y)
    return zhangdie


def Y_1(X, a, b):
    Y1 = []
    for i in range(0, len(X)):
        y = a * X[i] + b
        Y1.append(y)
    return Y1


def Y_2(Y, Y1):
    Y2 = np.array(Y) - np.array(Y1)
    return Y2


def CAR(A_coe, A_car, i_coe, i_car):
    a, b = Regression(A_coe, zhangdie(i_coe))
    Y = zhangdie(i_car)
    Y1 = Y_1(A_car, a, b)
    Y2 = Y_2(Y, Y1)

    return np.sum(Y2)


# zone = {(0,1),(0,2),(0,3),(0,4),(0,5),(0,6),(0,7),(0,8),(0,9),(0,10),
#         (-1,0),(-1,1),(-2,2),(-3,3),(-4,4),(-5,5),(-6,6),(-7,7),(-8,8),(-9,9),(-10,10)}


low = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, -2, -3, -4, -5, -6, -7, -8, -9, -10]
high = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
# print(len(low))
# print(len(high))

for i in range(0, len(low)):

    # save_name = str(i)+ '.xlsx'

    s = low[i]
    e = high[i]

    save_name = '(' + str(s) + ',' + str(e) + ')' + '.xlsx'
    # print(save_name)

    # 以1月20号为0，计算出需求时间区间内，表格对应的行数范围值
    s_x = -e + 12
    e_x = -s + 14

    A_coe_path = r'C:\Users\65487\Desktop\指数\中小板\19(coe)\指数行情.xlsx'
    A_car_path = r'C:\Users\65487\Desktop\指数\中小板\20(car)\指数行情.xlsx'
    coe_path = r'C:\Users\65487\Desktop\股票代码\19\中小板'
    car_path = r'C:\Users\65487\Desktop\股票代码\20\中小板'

    file_name = os.listdir(coe_path)

    wb = Workbook()
    ws = wb.active

    A_coe = zhangdie(load_price(A_coe_path, 30, 181, line='G'))
    A_car = zhangdie(load_price(A_car_path, s_x, e_x, line='G'))

    i = 1
    j = 1
    for f in file_name:
        i_coe = load_price(os.path.join(coe_path, f), 30, 181)
        i_car = load_price(os.path.join(car_path, f), s_x, e_x)
        # print(f)
        try:
            car = CAR(A_coe, A_car, i_coe, i_car)
            ws['A' + str(i)].value = Path(f).stem[:-4]
            ws['B' + str(i)].value = car
            i += 1
            # print(len(file_name)-i)
        except:
            # print(f, '数据异常')
            # print(i)
            ws['D' + str(j)].value = Path(f).stem[:-4]
            j += 1

    wb.save(r'C:\Users\65487\Desktop\result\中小板\30-181' + '/' + save_name)
    print(save_name + '已保存，有%d个公司未计算' % (j - 1))
    wb.close()
